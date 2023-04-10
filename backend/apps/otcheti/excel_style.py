from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle, PatternFill

ZAGOLOVOK1_EXCEL_STYLE = NamedStyle(
    name='zagolovok1_style',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='center',
                        vertical='top',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True, size=14),

)

ZAGOLOVOK1_BOTTOM_EXCEL_STYLE = NamedStyle(
    name='zagolovok1_bottom_style',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True, size=14),

)

SAPKA_EXCEL_STYLE = NamedStyle(
    name='sapka_style',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  ),

)

DATE_EXCEL_STYLE = NamedStyle(
    name='date_style',
    number_format='DD/MM/YYYY',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=False),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)

TEXT_EXCEL_STYLE_AND_BORDER = NamedStyle(
    name='text_style_and_border',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=False),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)

OFORMIT_KAK_TEKST = NamedStyle(
    name='oformit_kak_tekst',
    # fill=PatternFill(
    #     bgColor='00ff00'
    # ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True,),


)

OFORMIT_KAK_PRODLENIE = NamedStyle(
    name='oformit_kak_prodlenie',
    fill=PatternFill(
        start_color='00FF00',
        end_color='00FF00',
        fill_type='solid'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=False, size=10, name='Times New Roman'),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)


OFORMIT_KAK_PRODLENIE_BOLD = NamedStyle(
    name='oformit_kak_prodlenie_bold',
    fill=PatternFill(
        start_color='00FF00',
        end_color='00FF00',
        fill_type='solid'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True, size=10, name='Times New Roman'),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)

OFORMIT_KAK_NOVOE = NamedStyle(
    name='oformit_kak_novoe',
    fill=PatternFill(
        start_color='FFFF00',
        end_color='FFFF00',
        fill_type='solid'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=False, size=10, name='Times New Roman'),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)

OFORMIT_KAK_NOVOE_BOLD = NamedStyle(
    name='oformit_kak_novoe_bold',
    fill=PatternFill(
        start_color='FFFF00',
        end_color='FFFF00',
        fill_type='solid'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True, size=10, name='Times New Roman'),
    border=Border(left=Side(border_style='thin',
                            color='FF000000'),
                  right=Side(border_style='thin',
                             color='FF000000'),
                  top=Side(border_style='thin',
                           color='FF000000'),
                  bottom=Side(border_style='thin',
                              color='FF000000'),
                  diagonal=Side(border_style='thin',
                                color='FF000000'),
                  diagonal_direction=0,
                  outline=Side(border_style='thin',
                               color='FF000000'),
                  vertical=Side(border_style='thin',
                                color='FF000000'),
                  horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

)

TEXT_EXCEL_STYLE = NamedStyle(
    name='text_style',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=False),

)

TEXT_EXCEL_BOTTOM_STYLE = NamedStyle(
    name='text_style_bottom',
    fill=PatternFill(
        bgColor='FFFFFFFF'
    ),
    alignment=Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0),
    font=Font(bold=True),

)