import datetime
import os

from django.http import HttpResponseRedirect, HttpResponse
from docxtpl import DocxTemplate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from rest_framework.templatetags.rest_framework import data
from rest_framework.views import APIView

from apps.otcheti.excel_style import DATE_EXCEL_STYLE, SAPKA_EXCEL_STYLE, ZAGOLOVOK1_EXCEL_STYLE, \
    TEXT_EXCEL_BOTTOM_STYLE, ZAGOLOVOK1_BOTTOM_EXCEL_STYLE, TEXT_EXCEL_STYLE, TEXT_EXCEL_STYLE_AND_BORDER, \
    OFORMIT_KAK_TEKST
from apps.proekti.models import Proekt, ResheniePoProektu
from core.settings import BASE_DIR, MEDIA_ROOT


class RenderPovestkiApi(APIView):
    def get(self, request):
        reshenie = ResheniePoProektu.objects.filter(data_ispolnenia_po_resheniyu=datetime.date.today())

        wb = Workbook()
        wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title='Лист 1'),
        ws = ws[0]

        wb.add_named_style(DATE_EXCEL_STYLE)
        wb.add_named_style(SAPKA_EXCEL_STYLE)
        wb.add_named_style(OFORMIT_KAK_TEKST)
        wb.add_named_style(TEXT_EXCEL_BOTTOM_STYLE)
        wb.add_named_style(ZAGOLOVOK1_BOTTOM_EXCEL_STYLE)
        wb.add_named_style(TEXT_EXCEL_STYLE)
        wb.add_named_style(TEXT_EXCEL_STYLE_AND_BORDER)

        # настройки печати
        ws.print_title_rows = '1:1'  # the first row
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = '9'
        ws.page_margins.top = 0.75
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.bottom = 0.4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False

        ws.cell(row=1, column=1).value = f'Повестка совещания'
        ws.cell(row=2, column=1).value = f'г.Москва'
        ws.cell(row=2, column=9).value = f'{datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}'
        thin = Side(border_style="medium", color="000000")

        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:G2')
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 20
        ws.row_dimensions[3].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[1].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 40
        ws.row_dimensions[1].font = Font(bold=True)
        ws.row_dimensions[2].font = Font(bold=True)
        ws.row_dimensions[3].font = Font(bold=True)


        ws.cell(row=3, column=1).value = f'№ п/п'
        ws.cell(row=3, column=2).value = f'Адрес (адресные ориентиры)'
        ws.cell(row=3, column=3).value = 'Район города Москвы'
        ws.cell(row=3, column=4).value = 'Кадастровый номер земельного участка'
        ws.cell(row=3, column=5).value = '№ протокола'
        ws.cell(row=3, column=6).value = 'Дата первичного рассмотрения'
        ws.cell(row=3, column=7).value = 'Кол-во рассмотрений'
        ws.cell(row=3, column=8).value = 'Прошлое поручение'
        ws.cell(row=3, column=9).value = 'Новое поручение'
        ws.cell(row=3, column=10).value = 'Примечания'

        for i_row in range(3):
            for i_col in range(10):
                ws.cell(row=i_row+1, column=i_col+1).style = 'text_style_bottom'

        ws.cell(row=2, column=1).style = 'oformit_kak_tekst'

        # # ОФОРМЛЯЕМ ШАПКУ
        # for i in range(3):
        #     ws.cell(row=6, column=i + 1).style = 'sapka_style'

        index = 4
        for reshenie_i in reshenie:
            ws.row_dimensions[index].alignment = Alignment(horizontal="center", vertical="center")
            kol_vo_resheni = ResheniePoProektu.objects.filter(proekt_id=reshenie_i.proekt_id, dom_id=reshenie_i.dom.id).count()
            ws.row_dimensions[index].height = 35
            # Порядковый номер
            ws.cell(row=index, column=1).value = index - 3

            # Адрес
            ws.cell(row=index, column=2).value = f'{reshenie_i.dom.uliza}, {reshenie_i.dom.name}'

            # Район
            ws.cell(row=index, column=3).value = reshenie_i.dom.uliza.rayon.name


            ws.cell(row=index, column=4).value = index - 3

            # Номер протокола
            ws.cell(row=index, column=5).value = reshenie_i.nomer_protokola


            ws.cell(row=index, column=6).value = index - 3

            # Количество решений
            ws.cell(row=index, column=7).value = kol_vo_resheni

            # Прошлое поручение
            ws.cell(row=index, column=8).value = index - 3

            # Новое поручение
            ws.cell(row=index, column=9).value = reshenie_i.primechanie_po_resheniu

            # Примечание
            ws.cell(row=index, column=10).value = reshenie_i.dom.vladelez_zdania

            for i_col in range(10):
                ws.cell(row=index, column=i_col+1).style = 'text_style_and_border'
            index += 1

        os_dir_path = os.path.join(MEDIA_ROOT, 'otcheti')
        if not os.path.exists(os_dir_path):
            os.makedirs(os_dir_path)

        excel_name = f'povestka-{datetime.date.today()}.xlsx'
        excel_path = os.path.join(os_dir_path, excel_name)

        wb.save(filename=excel_path)
        return HttpResponseRedirect(f'/media/otcheti/{excel_name}')